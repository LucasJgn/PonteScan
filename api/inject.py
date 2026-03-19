from http.server import BaseHTTPRequestHandler
import json
import base64
import io
import openpyxl

class handler(BaseHTTPRequestHandler):

    def do_OPTIONS(self):
        self.send_response(200)
        self._send_cors()
        self.end_headers()

    def do_POST(self):
        content_length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(content_length)

        try:
            data = json.loads(body)
            file_b64 = data.get('fileBase64')
            values = data.get('values', {})
            target_age = data.get('age')

            if not file_b64 or target_age is None:
                self._respond(400, {'error': 'fileBase64 et age requis'})
                return

            # Décoder le fichier Excel
            file_bytes = base64.b64decode(file_b64)
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes))

            # Trouver l'onglet "Fiche technique de production" ou premier onglet
            sheet_name = None
            for name in wb.sheetnames:
                if 'fiche' in name.lower() or 'production' in name.lower():
                    sheet_name = name
                    break
            if not sheet_name:
                sheet_name = wb.sheetnames[0]

            ws = wb[sheet_name]

            # Chercher la ligne où colonne C = age
            target_row = None
            for row in ws.iter_rows():
                cell_c = row[2].value  # colonne C = index 2
                if cell_c is not None:
                    try:
                        if int(cell_c) == int(target_age):
                            target_row = row[0].row
                            break
                    except (ValueError, TypeError):
                        continue

            if target_row is None:
                self._respond(404, {'error': f'Semaine âge {target_age} introuvable'})
                return

            # Mapping colonnes → valeurs
            col_map = {
                'D': values.get('mortalite'),
                'G': values.get('effectif_fin'),
                'H': values.get('ponte_j1'),
                'I': values.get('ponte_j2'),
                'J': values.get('ponte_j3'),
                'K': values.get('ponte_j4'),
                'L': values.get('ponte_j5'),
                'M': values.get('ponte_j6'),
                'N': values.get('ponte_j7'),
                'P': values.get('ponte_hebdo'),
                'W': values.get('declasses'),
                'AM': values.get('conso_aliment_kg'),
                'AY': values.get('conso_eau_L'),
            }
            # Optionnel
            if values.get('poids_oeufs'):
                col_map['AB'] = values['poids_oeufs']
            if values.get('poids_vif'):
                col_map['AV'] = values['poids_vif']
            if values.get('homogeneite'):
                col_map['AW'] = values['homogeneite']
            if values.get('observations'):
                col_map['BF'] = values['observations']

            # Injecter — openpyxl préserve les styles existants
            for col, val in col_map.items():
                if val is not None and val != '' and val != 0:
                    ws[f'{col}{target_row}'] = val

            # Sauvegarder en mémoire
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            result_b64 = base64.b64encode(output.read()).decode('utf-8')

            self._respond(200, {
                'fileBase64': result_b64,
                'rowFound': target_row,
                'sheetName': sheet_name
            })

        except Exception as e:
            self._respond(500, {'error': str(e)})

    def _send_cors(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')

    def _respond(self, status, data):
        self.send_response(status)
        self._send_cors()
        self.send_header('Content-Type', 'application/json')
        self.end_headers()
        self.wfile.write(json.dumps(data).encode())
